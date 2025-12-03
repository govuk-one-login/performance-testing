#!/usr/bin/env python3
"""
Generate Excel report from performance test data
Usage: python3 generate_excel_report.py <timestamp> <scenario1> <scenario2> ...
"""

import sys
import csv
import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Install with: pip3 install openpyxl")
    sys.exit(1)


def create_excel_report(timestamp, scenarios, data_file):
    """Create Excel workbook with multiple tabs"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Read the data file
    with open(data_file, 'r') as f:
        data = json.load(f)

    # Create tabs
    create_response_time_tab(wb, scenarios, data)
    create_journey_error_tab(wb, data)
    create_http_requests_tab(wb, scenarios, data)

    # Save workbook
    filename = f"Performance_Report_{timestamp}.xlsx"
    wb.save(filename)
    print(f"📊 Excel report generated: {filename}")
    return filename


def create_response_time_tab(wb, scenarios, data):
    """Create Response Time tab with all scenario tables"""
    ws = wb.create_sheet("Response Time")

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    for scenario in scenarios:
        # Scenario header
        ws.cell(row, 1, f"SCENARIO: {scenario}").font = Font(bold=True, size=14)
        row += 2

        # Iterations Started
        ws.cell(row, 1, "ITERATIONS STARTED (STEADY-STATE)").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row, 1, "Scenario").font = header_font
        ws.cell(row, 1).fill = header_fill
        ws.cell(row, 2, "Count").font = header_font
        ws.cell(row, 2).fill = header_fill
        row += 1
        ws.cell(row, 1, scenario)
        ws.cell(row, 2, data['scenarios'][scenario]['iter_started'])
        row += 2

        # Iterations Completed & Throughput
        ws.cell(row, 1, "ITERATIONS COMPLETED & THROUGHPUT (STEADY-STATE)").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row, 1, "Scenario").font = header_font
        ws.cell(row, 1).fill = header_fill
        ws.cell(row, 2, "Count").font = header_font
        ws.cell(row, 2).fill = header_fill
        ws.cell(row, 3, "Throughput/sec").font = header_font
        ws.cell(row, 3).fill = header_fill
        row += 1
        ws.cell(row, 1, scenario)
        ws.cell(row, 2, data['scenarios'][scenario]['iter_completed'])
        ws.cell(row, 3, data['scenarios'][scenario]['throughput'])
        row += 2

        # Steady-State Percentiles
        ws.cell(row, 1, "STEADY-STATE PERCENTILES").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row, 1, "Transaction Name").font = header_font
        ws.cell(row, 1).fill = header_fill
        ws.cell(row, 2, "P95(ms)").font = header_font
        ws.cell(row, 2).fill = header_fill
        ws.cell(row, 3, "P99(ms)").font = header_font
        ws.cell(row, 3).fill = header_fill
        ws.cell(row, 4, "Count").font = header_font
        ws.cell(row, 4).fill = header_fill
        row += 1

        # Read CSV file for this scenario
        csv_file = f"SS_RT_{scenario}_{data['timestamp']}.csv"
        if Path(csv_file).exists():
            with open(csv_file, 'r') as f:
                reader = csv.DictReader(f)
                for csv_row in reader:
                    ws.cell(row, 1, csv_row['Transaction Name'])
                    ws.cell(row, 2, float(csv_row['P95(ms)']))
                    ws.cell(row, 3, float(csv_row['P99(ms)']))
                    ws.cell(row, 4, int(csv_row['Count']))
                    row += 1

        row += 2

        # Full Test Duration Percentiles
        ws.cell(row, 1, "FULL TEST DURATION PERCENTILES").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row, 1, "Transaction Name").font = header_font
        ws.cell(row, 1).fill = header_fill
        ws.cell(row, 2, "P95(ms)").font = header_font
        ws.cell(row, 2).fill = header_fill
        ws.cell(row, 3, "P99(ms)").font = header_font
        ws.cell(row, 3).fill = header_fill
        ws.cell(row, 4, "Count").font = header_font
        ws.cell(row, 4).fill = header_fill
        row += 1

        # Add full duration percentiles data
        if 'full_duration_percentiles' in data and scenario in data['full_duration_percentiles']:
            for item in data['full_duration_percentiles'][scenario]:
                ws.cell(row, 1, item['transaction'])
                ws.cell(row, 2, float(item['p95']))
                ws.cell(row, 3, float(item['p99']))
                ws.cell(row, 4, int(item['count']))
                row += 1

        row += 3  # Spacing between scenarios

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_journey_error_tab(wb, data):
    """Create Journey Error Rate tab"""
    ws = wb.create_sheet("Journey Error Rate")

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    row = 1
    ws.cell(row, 1, "JOURNEY ERROR RATE (ENTIRE TEST)").font = Font(bold=True, size=14)
    row += 2

    # Headers
    headers = ["Scenario", "Started", "Completed", "Failed", "Error Rate%"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    # Data rows
    for scenario, scenario_data in data['scenarios'].items():
        ws.cell(row, 1, scenario)
        ws.cell(row, 2, scenario_data['total_started'])
        ws.cell(row, 3, scenario_data['total_completed'])
        ws.cell(row, 4, scenario_data['failed'])
        ws.cell(row, 5, scenario_data['error_rate'])
        row += 1

    # Total row
    ws.cell(row, 1, "TOTAL").font = Font(bold=True)
    ws.cell(row, 2, data['total']['started']).font = Font(bold=True)
    ws.cell(row, 3, data['total']['completed']).font = Font(bold=True)
    ws.cell(row, 4, data['total']['failed']).font = Font(bold=True)
    ws.cell(row, 5, data['total']['error_rate']).font = Font(bold=True)

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


def create_http_requests_tab(wb, scenarios, data):
    """Create HTTP Requests tab with all scenario tables"""
    ws = wb.create_sheet("HTTP Requests")

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    row = 1

    for scenario in scenarios:
        ws.cell(row, 1, f"SCENARIO: {scenario}").font = Font(bold=True, size=14)
        row += 1
        ws.cell(row, 1, "HTTP REQUESTS BY STATUS (ENTIRE TEST)").font = Font(bold=True, size=12)
        row += 1

        # Headers
        ws.cell(row, 1, "Group Name").font = header_font
        ws.cell(row, 1).fill = header_fill
        ws.cell(row, 2, "Status").font = header_font
        ws.cell(row, 2).fill = header_fill
        ws.cell(row, 3, "Count").font = header_font
        ws.cell(row, 3).fill = header_fill
        row += 1

        # Data rows
        if scenario in data['http_requests']:
            for req in data['http_requests'][scenario]:
                ws.cell(row, 1, req['group'])
                ws.cell(row, 2, req['status'])
                ws.cell(row, 3, req['count'])
                row += 1

        row += 3  # Spacing between scenarios

    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 generate_excel_report.py <timestamp> <data_file>")
        sys.exit(1)

    timestamp = sys.argv[1]
    data_file = sys.argv[2]

    # Read scenarios from data file
    with open(data_file, 'r') as f:
        data = json.load(f)

    scenarios = list(data['scenarios'].keys())

    create_excel_report(timestamp, scenarios, data_file)
